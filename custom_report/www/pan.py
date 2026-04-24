import frappe
import json
import csv
import io
from custom_report.db_connection import get_dr_connection

@frappe.whitelist(allow_guest=True)
def get_pan_details(pan_no=None):
    pan_no = pan_no or frappe.form_dict.get('pan_no')
    if not pan_no: return {"success": False, "error": "PAN number missing"}
    conn = None
    try:
        try: conn = get_dr_connection()
        except Exception as e: return {"success": False, "error": f"DB Connection Failed: {str(e)}"}
        cursor = conn.cursor()
        query = """
            SELECT distinct 
                e.referencenumber, g.cif_id, g.foracid, g.acct_name, a.cust_dob,
                g.schm_code, g.schm_type, g.acct_opn_date, g.sol_id, c.phonenolocalcode,
                s.division_name, s.circle_office_name, s.sol_desc
            FROM tbaadm.gam g, crmuser.cphone c, tbaadm.sol s, crmuser.entitydocument e, crmuser.accounts a
            WHERE c.preferredflag='Y' AND g.cif_id=c.phone_b2kid AND g.sol_id=s.sol_id 
            AND g.cif_id = a.orgkey AND e.orgkey=g.cif_id AND e.docdescr='PAN CARD'
            AND g.schm_code IN ('1001','1002','1003') AND g.entity_cre_flg='Y'
            AND e.referencenumber = %s
        """
        cursor.execute(query, (pan_no,))
        r = cursor.fetchone()
        if r:
            return {
                "success": True, 
                "data": {
                    "pan":r[0],"cif":r[1],"acc":r[2],"name":r[3],"dob":str(r[4]),
                    "sc":r[5],"st":r[6],"od":str(r[7]),"sol":r[8],"ph":r[9],
                    "dist":r[10],"zone":r[11],"sd":r[12]
                }
            }
        return {"success": False, "error": "No records found."}
    except Exception as e: return {"success": False, "error": str(e)}
    finally:
        if conn: conn.close()

@frappe.whitelist(allow_guest=True)
def get_pan_list_from_file():
    """Extract PAN numbers from uploaded file (CSV/Excel)"""
    # Try multiple ways to get the file to avoid 400 Bad Request
    file_obj = None
    if frappe.request.files and 'file' in frappe.request.files:
        file_obj = frappe.request.files['file']
    elif frappe.form_dict.get('file'):
        # In some cases Frappe puts it here
        file_obj = frappe.form_dict.get('file')

    if not file_obj:
        return {"success": False, "error": "No file received. Please ensure you are uploading a valid CSV or Excel file."}
    filename = file_obj.filename.lower()
    pan_list = []
    
    try:
        if filename.endswith('.xlsx'):
            from openpyxl import load_workbook
            content = file_obj.read()
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                if row and row[0]:
                    pan_list.append(str(row[0]).strip())
        else:
            # CSV Handling
            content = file_obj.read().decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                if row and row[0]:
                    pan_list.append(str(row[0]).strip())

        if not pan_list:
            return {"success": False, "error": "File is empty."}

        # Filter out headers
        headers = ['pan', 'referencenumber', 'reference no', 'reference_no', 'pan no', 'pan number']
        if pan_list[0].lower() in headers:
            pan_list = pan_list[1:]

        # Deduplicate and return
        return {"success": True, "pans": list(dict.fromkeys(pan_list))}
        
    except Exception as e:
        return {"success": False, "error": f"File processing error: {str(e)}"}

@frappe.whitelist(allow_guest=True)
def fetch_pan_batch(pan_json):
    """Fetch 13 fields for a batch of 100 PANs"""
    try:
        pans = json.loads(pan_json)
    except:
        return {"success": False, "error": "Invalid JSON data"}
        
    if not pans: return {"success": True, "results": []}
    
    results = []
    conn = None
    try:
        conn = get_dr_connection()
        cursor = conn.cursor()
        
        placeholders = ','.join(['%s'] * len(pans))
        q = f"""
            SELECT distinct 
                e.referencenumber, g.cif_id, g.foracid, g.acct_name, a.cust_dob,
                g.schm_code, g.schm_type, g.acct_opn_date, g.sol_id, c.phonenolocalcode,
                s.division_name, s.circle_office_name, s.sol_desc
            FROM tbaadm.gam g, crmuser.cphone c, tbaadm.sol s, crmuser.entitydocument e, crmuser.accounts a
            WHERE c.preferredflag='Y' AND g.cif_id=c.phone_b2kid AND g.sol_id=s.sol_id 
            AND g.cif_id = a.orgkey AND e.orgkey=g.cif_id AND e.docdescr='PAN CARD'
            AND g.schm_code IN ('1001','1002','1003') AND g.entity_cre_flg='Y'
            AND e.referencenumber IN ({placeholders})
        """
        cursor.execute(q, tuple(pans))
        db_rows = cursor.fetchall()
        
        found_map = {r[0]: r for r in db_rows}
        for p in pans:
            if p in found_map:
                r = found_map[p]
                results.append({
                    "pan":r[0], "name":r[3], "cif":r[1], "acc":r[2], "dob":str(r[4]),
                    "ph":r[9], "sc":r[5], "st":r[6], "od":str(r[7]), "sol":r[8],
                    "dist":r[10], "zone":r[11], "sd":r[12], "status":"Found"
                })
            else:
                results.append({"pan":p, "status":"Not Found"})
                
        return {"success": True, "results": results}
    except Exception as e:
        return {"success": False, "error": f"DB Error: {str(e)}"}
    finally:
        if conn: conn.close()
